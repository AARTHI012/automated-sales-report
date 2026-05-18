import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
import yaml

def generate_excel_report(df, summary, config):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    
    # Header
    ws['A1'] = config['report']['title']
    ws['A2'] = f"Period: {config['report']['period']} | Generated: {datetime.now().strftime('%Y-%m-%d')}"
    
    # Summary Stats
    row = 4
    for key, value in summary.items():
        ws[f'A{row}'] = key.replace('_', ' ').title()
        ws[f'B{row}'] = value
        row += 1
    
    # Raw Data
    wb.create_sheet("Raw Data")
    for r in dataframe_to_rows(df, index=False, header=True):
        wb["Raw Data"].append(r)
    
    # Monthly Summary
    wb.create_sheet("Monthly Summary")
    monthly = df.groupby('Month').sum(numeric_only=True)
    for r in dataframe_to_rows(monthly.reset_index(), index=False, header=True):
        wb["Monthly Summary"].append(r)
    
    wb.save(config['output']['filename'].format(date=datetime.now().strftime("%Y%m%d")))
    print(f"✅ Report generated: {config['output']['filename']}")
